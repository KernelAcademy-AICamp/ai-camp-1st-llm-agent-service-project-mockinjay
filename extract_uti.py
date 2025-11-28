#!/usr/bin/env python3
import sys
import json

try:
    import openpyxl
except ImportError:
    print("openpyxl not installed, trying to install...")
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    import openpyxl

wb = openpyxl.load_workbook('test.xlsx')
ws = wb.active

# Get header row
headers = []
for cell in ws[1]:
    headers.append(cell.value)

# Find rows with UTI-001 to UTI-005
target_ids = ['UTI-001', 'UTI-002', 'UTI-003', 'UTI-004', 'UTI-005']
results = []

for row in ws.iter_rows(min_row=2, values_only=True):
    if row[0] in target_ids:
        row_dict = {}
        for i, header in enumerate(headers):
            if i < len(row):
                row_dict[header] = row[i]
        results.append(row_dict)

# Print results in a readable format
for result in results:
    print("=" * 80)
    for key, value in result.items():
        if value is not None:
            print(f"{key}: {value}")
    print()
